import openpyxl
import urllib.request
import urllib.parse
import re


def readExfile(filepath):
    inwb = openpyxl.load_workbook(filepath)  # 读文件
    # sourceSheet = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    # ws = inwb.get_sheet_by_name(sourceSheet[0])  # 获取第一个sheet内容
    sheet1=inwb.sheetnames
    ws=inwb[sheet1[0]]
    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    print("行数" +str(rows))
    for r in range(2, rows):
        for c in range(1, cols):
            for c in (1,2,4,5,6):
                pass
                # line=ws.cell(r, c)
                id= str(ws.cell(r,1).value)
                gender=str(ws.cell(r,2).value)
                age=str(ws.cell(r,4).value)
                dia=ws.cell(r,5).value
                dis=str(re.sub(';','\",\"',dia))
                drug=str(re.sub(",","\",\"",ws.cell(r,6).value))
        data='{"emid":"'+id+'","gender":'+gender+',"age":'+age+',"diagnose":["'+dis+'"],"drugNames":["'+drug+']"}'
        return data
        # if r == 5:
        #     break

def posturl(url,file):
    url='http://171.cdfortis.com/drugAssistant/recipeCheck'
    errorrequest=[]
    data=readExfile(file)
    for i in data:
        data = urllib.parse.urlencode(data)
        data = data.encode('utf-8')
        response = urllib.request.urlopen(url=url, data=data)
        res = response.read()
        if re.search("完成",res):
            pass
        else:
            errorrequest.append(data)
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for row in range(1, 70000):
        for col in range(1, 4):
            outws.cell(row, col).value = row * 2  # 写文件
        print(row)
    saveExcel = "E:\\pydemo\\exceldata\\testresult.xlsx"
    outwb.save(saveExcel)  # 一定要记得保存


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    readExfile('E:/pydemo/exceldata/饿了么t.xlsx')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
